# -*- coding: utf-8 -*-
from time import gmtime, strftime
import hexdump
import argparse
import subprocess
from argparse import RawTextHelpFormatter
import os
import fnmatch
import plistlib
import biplist
import pandas as pd
import numpy as np
import ccl_bplist
import re
from config import CONFIG
import IO_databaseOperations as db
import logging
logger = logging.getLogger('root')
import datetime
from openpyxl import Workbook, load_workbook
from struct import unpack
import time
import IO_blobOperations
import IO_browserOperations as browser
import PARSER_download as download

import sys  
reload(sys)  
sys.setdefaultencoding('utf8')

#declare global variables
usbmsc = pd.DataFrame()
timestamp = str(datetime.datetime.strftime(datetime.datetime.today(),'%Y%m%d%H%M%S'))
entrytype = {'8': 'Time Machine (AFPFS), AFP File Shares, OSXFUSE Volumes', '16':'Network Hard Drive, iDisk, Computer', '128':'iDisk', '261': 'Hard Drive, Boot Hard Drive', '515' : 'USB Flash, Time Machine Backups, Disk Image (HFS, MBR)', '517':'USB Hard Drive (FAT/ExFAT/HFS+)', '1024':'Remote Disk', '1027':'Disk Image (Bzip, VAX COFF Executable), DVD', '1029':'External HDD (NTFS)'}

def ParseSFL(MRUFile):
	sfldf = pd.DataFrame()
	if "com.apple.LSSharedFileList.ApplicationRecentDocuments" in MRUFile:
		sfl_columns=['order', 'name', 'url', 'creationdate', 'volume', 'packagename','user']
	else:
		sfl_columns=['order', 'name', 'url', 'creationdate', 'volume','user']
	sflitem = {}
	try:
		plistfile = open(MRUFile, "rb")
		plist = ccl_bplist.load(plistfile)
		plist_objects = ccl_bplist.deserialise_NsKeyedArchiver(plist, parse_whole_structure=True)

		if plist_objects["root"]["NS.objects"][1]["NS.keys"][0] == "com.apple.LSSharedFileList.MaxAmount":
			numberOfItems = plist_objects["root"]["NS.objects"][1]["NS.objects"][0]
			print "Max number of recent items in this plist: " + str(numberOfItems)

		if plist_objects["root"]["NS.keys"][2] == "items":
			items = plist_objects["root"]["NS.objects"][2]["NS.objects"] 
			for n,item in enumerate(items):
				print list(item.keys())
				sflitem['order'] = str(item["order"])
				try:
					sflitem['name'] = item["name"]
					print item["name"]
				except:
					sflitem['name'] = ""
				sflitem['url'] = item["URL"]['NS.relative']
				try:
					data = item["bookmark"]
					try:
						sflitem['creationdate'] = IO_blobOperations.extract_creationdate(item["bookmark"]['NS.data'])
						sflitem['volume'] = IO_blobOperations.extract_volume(item["bookmark"]['NS.data'])
					except:
						sflitem['creationdate'] = IO_blobOperations.extract_creationdate(item["bookmark"])
						sflitem['volume'] = IO_blobOperations.extract_volume(item["bookmark"])
				except:
					sflitem['creationdate'] = "Not set"
					sflitem['volume'] = "Not set"
				if "com.apple.LSSharedFileList.ApplicationRecentDocuments" in MRUFile:
					sflitem['packagename']=os.path.basename(MRUFile).replace('.sfl','')
				sflitem['user'] = re.search(r'([^\\]*)_.*\.sfl',MRUFile).group(1)
				sfldf = sfldf.append(pd.DataFrame(sflitem, index=['i',], columns=sfl_columns), ignore_index=True)
	except Exception as e:
		print "Cannot open file: " + MRUFile + str(e)
	return sfldf

def ParseSidebarPlist(sidebarfile):
	sidebardf = pd.DataFrame()
	sidebar_columns=['itemno', 'itemname', 'voluuid', 'entrytype']
	sidebaritem = {}
	try:
		plist = biplist.readPlist(sidebarfile)
		for n,item in enumerate(plist["systemitems"]["VolumesList"]):
			itemlist = list(item.keys())
			print "    [Item Number: " + str(n) + "] '" + item["Name"] + "'"
			sidebaritem['itemno'] = str(n)
			sidebaritem['itemname'] = item["Name"]
			if "Bookmark" in itemlist:
				sidebaritem['voluuid'] = IO_blobOperations.extract_voluuid(item["Bookmark"])
			else:
				if "Alias" in itemlist:
					sidebaritem['voluuid'] = IO_blobOperations.extract_voluuid(item["Alias"])
				else:
					sidebaritem['voluuid'] = "Not set"
			sidebaritem['entrytype'] = entrytype[str(item["EntryType"])]  
			sidebardf = sidebardf.append(pd.DataFrame(sidebaritem, index=['i',], columns=sidebar_columns), ignore_index=True)
		return sidebardf
	except Exception as e:
		print "Cannot open file: " + sidebarfile + str(e)

def ParseFinderPlist(MRUFile):
	finderdf = pd.DataFrame()
	finder_columns=['itemno', 'itemname', 'creationdate','volume','user']
	finderitem = {}
	try:
		plist = biplist.readPlist(MRUFile)
		for n,item in enumerate(plist["FXRecentFolders"]):
			finderitem['itemno'] = str(n)
			finderitem['itemname'] = item["name"]
			finderitem['creationdate'] = IO_blobOperations.extract_creationdate(item["file-bookmark"])  
			finderitem['volume'] = IO_blobOperations.extract_volume(item["file-bookmark"])  
			finderitem['user'] = re.search(r'([^\\]*)_com\.apple\.finder\.plist',MRUFile).group(1)
			finderdf = finderdf.append(pd.DataFrame(finderitem, index=['i',], columns=finder_columns), ignore_index=True)
	except Exception as e:
		print "Cannot open file: " + MRUFile + str(e)
	return finderdf

def ParseInstallHistory(MRUFile):
	installdf = pd.DataFrame()
	install_columns=['itemno', 'itemname', 'date', 'version', 'processname']
	installitem = {}
	try:
		plist=plistlib.readPlist(MRUFile)
		for n,item in enumerate(plist):
			installitem['itemno'] = str(n)
			installitem['itemname'] = item["displayName"]
			installitem['date'] = item["date"]
			installitem['version'] = item["displayVersion"]
			installitem['processname'] = item["processName"]
			installdf = installdf.append(pd.DataFrame(installitem, index=['i',], columns=install_columns), ignore_index=True)
	except Exception as e:
		print "Cannot open file: " + MRUFile + str(e)
	return installdf

def ParseSafariWebHistory(historyFile, owner):
	safarihistorydf = pd.DataFrame()
	safarihistory_columns=['URL', 'name', 'firstvisitdate','profile','browser']
	safarihistoryitem = {}
	try:
		plist = biplist.readPlist(historyFile)
		print plist.keys()
		safarihistoryitem['URL'] = plist['URL']
		safarihistoryitem['name'] = plist['Name']
		safarihistoryitem['firstvisitdate'] = ""
		safarihistoryitem['profile'] = owner
		safarihistoryitem['browser'] = "Safari"
		safarihistorydf = safarihistorydf.append(pd.DataFrame(safarihistoryitem, index=['i',], columns=safarihistory_columns), ignore_index=True)
	except Exception as e:
		print "Cannot open file: " + str(e)
	return safarihistorydf

def ParseSystemUSB(file, databaseConnectionHandle):
	usbdf = pd.DataFrame() 
	usb = {}
	usb_columnns = ['loggedtime', 'serialnumber','vendorid','vendorname','productid','productname','versionnumber']
	#usb = { 'serialnumber' : '', 'vendorid' : '', 'vendorname' : '', 'productid' : '', 'productname': '','versionnumber' : '', 'loggedtime' : ''}
	with open(file) as f:
		data = f.read().splitlines()
		for lines in data:
			if "USBMSC" in lines:
				#Full Message
				usbmsc_data = lines.split(": ")
				#SNO, VID, PID, VersionNO
				usbmsc_data_2 = usbmsc_data[-1].split(",")[0].split( )
				usb['loggedtime'] = re.match(r'.*\d{,2}:\d{,2}\d{,2}' ,usbmsc_data[0]).group(0)
				usb['serialnumber'] = usbmsc_data_2[0]
				usb['vendorid'] = usbmsc_data_2[1].replace("0x", "").zfill(4)
				#Look up to database
				cur = databaseConnectionHandle.cursor()
				query = "SELECT vendorname FROM usb_id.vendor_details WHERE vendorid = %s"
				logger.info("query is : " + str(query))
				cur.execute(query,(str(usb['vendorid']),))
				result = cur.fetchall()
				for i in result:        
					usb['vendorname'] = i[0]
				usb['productid'] = usbmsc_data_2[2].replace("0x", "").zfill(4)
				query = "SELECT prodname FROM usb_id.product_details WHERE vendorid = %s AND prodid = %s"
				logger.info("query is : " + str(query))
				cur.execute(query,(str(usb['vendorid']),str(usb['productid'])))
				result = cur.fetchall()
				for i in result:        
					usb['productname'] = i[0]
				usb['versionnumber'] = usbmsc_data_2[3].replace("0x", "")
				usbdf = usbdf.append(pd.DataFrame(usb, index=['i',], columns=usb_columnns), ignore_index=True)
		return usbdf

def outputToExcel(data, worksheet, directory, projectname, column="", checkrow=True):
	workbook = load_workbook(filename= './Results/' + projectname + '/' + imgname + '-Summary-' + timestamp + '.xlsx')
	writer = pd.ExcelWriter('./Results/' + projectname + '/' + imgname + '-Summary-' + timestamp + '.xlsx', engine='openpyxl')
	writer.book = workbook
	writer.sheets = dict((ws.title,ws) for ws in workbook.worksheets)

	if checkrow:
		lastRow = workbook[worksheet].max_row
	else:
		lastRow = 2
	if column == "":
		data.to_excel(writer,sheet_name=worksheet,index=False,header=False,startrow=lastRow)
	else:
		data.to_excel(writer,sheet_name=worksheet,startcol=column, index=False,header=False,startrow=lastRow)
	writer.save()

#Miscellaneous Functions
def clean(df, columns):
	for col in df.select_dtypes([np.object]).columns[1:]:
		df[col] = df[col].str.replace('[\000-\010]|[\013-\014]|[\016-\037]', '')
	return df

def ExtractGZ(file, root):
	subprocess.call(['.\\Tools\\7-Zip\\7z.exe','e', file, '-o'+root])

# def ParseSystemLogUSB

if __name__ == "__main__":
	#Image name is obtained from Incident Log.txt AND/OR *-log.txt from memory
	# DATABASE = CONFIG['DATABASE']
	# dbhandle = db.databaseConnect(DATABASE['HOST'], DATABASE['DATABASENAME'], DATABASE['USER'], DATABASE['PASSWORD'])
	# logger.info("dbhandle is " + str(dbhandle))

	parser = argparse.ArgumentParser(
		description='\
	Parse the Mac MRU (Most Recently Used) Plist Files \
	\n\n\tMac MRU File Locations: \
	\n\t- /Users/<username>/Library/Preferences/<bundle_id>.LSShardFileList.plist\
	\n\t- /Users/<username>/Library/Preferences/com.apple.finder.plist\
	\n\t- [10.10-] /Users/<username>/Library/Preferences/com.apple.recentitems.plist\
	\n\t- [10.11+] /Users/<username>/Library/Library/Application Support/com.apple.sharedfilelist/com.apple.LSSharedFileList.ApplicationRecentDocuments/<bundle_id>.sfl\
	\n\t- [10.11+] /Users/<username>/Library/Library/Application Support/com.apple.sharedfilelist/RecentApplications.sfl\
	\n\t- [10.11+] /Users/<username>/Library/Library/Application Support/com.apple.sharedfilelist/RecentDocuments.sfl\
	\n\t- [10.11+] /Users/<username>/Library/Library/Application Support/com.apple.sharedfilelist/RecentServers.sfl\
	\n\t- [10.11+] /Users/<username>/Library/Library/Application Support/com.apple.sharedfilelist/RecentHosts.sfl\
	\n\t- MS Office 2011 - /Users/<username>/Library/Preferences/com.microsoft.office.plist\
	\n\t- MS Office 2016 - /Users/<username>/Library/Containers/com.microsoft.<app>/Data/Library/Preferences/com.microsoft.<app>.securebookmarks.plist \
	\n \
	\n\tVersion: 1.1\
	\n\tUpdated: 08/15/2016\
	\n\tAuthor: Sarah Edwards | @iamevltwin | mac4n6.com | oompa@csh.rit.edu\
	\n\
	\n\tDependencies:\
	\n\t\thexdump.py: https://pypi.python.org/pypi/hexdump\
	\n\t\tccl_bplist.py: https://github.com/cclgroupltd/ccl-bplist'
		, prog='macMRU.py'
		, formatter_class=RawTextHelpFormatter)
	parser.add_argument('-d', dest='directory', required=True, type=str, help="Directory containing evidence files")
	parser.add_argument('-p', dest='project', required=True, type=str, help="Project Name")
	# parser.add_argument('--blob', action='store_true', help="Include hex dump of Bookmark BLOBs in standard output (can very ver")
	args = parser.parse_args()
	directory = args.directory

	#Check if results folder exist, if not, create it.
	resultsDir = os.getcwd() + "/Results"
	if not os.path.exists(resultsDir):
		try:
			os.makedirs(resultsDir)
		except:
			logging.error("Unable to create results folder")
			sys.exit()

	projResultsDir = os.getcwd() + "/Results/" + args.project 
	if not os.path.exists(projResultsDir):
		try:
			os.makedirs(projResultsDir)
		except:
			logging.error("Unable to create Project results folder")
			sys.exit()

	imgname = os.path.split(args.directory)[1]
	workbook = load_workbook(filename= 'LOGAN_Host_Analysis_Checklist_Results_template.xlsx')
	writer = pd.ExcelWriter('./Results/' + args.project + '/' + imgname + '-Summary-' + timestamp + '.xlsx', engine='openpyxl')
	writer.book = workbook
	writer.sheets = dict((ws.title,ws) for ws in workbook.worksheets)
	writer.save()

	print "###### MacMRU Parser v1.1 ######"

	#User-specific data (Merge all user data into one variable before output to excel)
	finder_Data = pd.DataFrame()
	recentapp_Data = pd.DataFrame() 
	recentdocs_Data = pd.DataFrame() 
	recenthosts_Data = pd.DataFrame() 
	recentservers_Data = pd.DataFrame()
	apprecentdocs_Data = pd.DataFrame() 
	browserHistory = pd.DataFrame() 
	browserDownload = pd.DataFrame()

	for root, dirs, filenames in os.walk(unicode(directory, 'utf-8')):
		for f in filenames:
			file = os.path.join(root,f)
#USB OUTPUT
#----------
			# if "system.log" in file:
			#     if not file.endswith(".gz"):
			#         USBMSC_Data=ParseSystemUSB(file, dbhandle)
			#         outputToExcel(USBMSC_Data, "USB", args.directory, args.project)
			#     # else:
			#         # print "=============================================================================="
			#         # print "Extracting file: " + file
			#         # ExtractGZ(file,root)
			#         # print "=============================================================================="
			#         # print "Processing file: " + file
			#         # ParseSystemUSB(re.sub(r'\.gz$', '', file, flags=re.IGNORECASE))
#             if "sidebarlists.plist" in file:
#                 sidebar_Data = ParseSidebarPlist(file)
#                 sidebar_Data = clean(sidebar_Data, list(sidebar_Data))
#                 outputToExcel(sidebar_Data, "USB", args.directory, args.project, 7, False)
# #INSTALLED SOFTWARE OUTPUT
# #-------------------------
#             # if "InstallHistory.plist" in file:
#             #     install_Data = ParseInstallHistory(file)
#             #     install_Data = clean(install_Data, list(install_Data))
#             #     print install_Data
#             #     outputToExcel(install_Data, "Installed_software", args.directory, args.project)
# #RECENT FILES OUTPUT
# #-------------------------
#             if "Recent Items" in file:
#                 if "com.apple.LSSharedFileList.ApplicationRecentDocuments" in file:
#                     data = ParseSFL(file)
#                     apprecentdocs_Data = apprecentdocs_Data.append(data, ignore_index=True)
#                     apprecentdocs_Data = clean(apprecentdocs_Data, list(apprecentdocs_Data))
#                 if "RecentHosts.sfl" in file:
#                     data = ParseSFL(file)
#                     recenthosts_Data = recenthosts_Data.append(data, ignore_index=True)
#                     recenthosts_Data = clean(recenthosts_Data, list(recenthosts_Data))
#                 if "RecentServers.sfl" in file:
#                     data = ParseSFL(file)
#                     recentservers_Data = recentservers_Data.append(data, ignore_index=True)
#                     recentservers_Data = clean(recentservers_Data, list(recentservers_Data))
#                 if "RecentDocuments.sfl" in file:
#                     data = ParseSFL(file)
#                     recentdocs_Data = recentdocs_Data.append(data, ignore_index=True)
#                     recentdocs_Data = clean(recentdocs_Data, list(recentdocs_Data))
#                 if "RecentApplications.sfl" in file:
#                     data = ParseSFL(file)
#                     recentapp_Data = recentapp_Data.append(data, ignore_index=True)
#                     recentapp_Data = clean(recentapp_Data, list(recentapp_Data))
#                 if "com.apple.finder.plist" in file:
#                     data = ParseFinderPlist(file)
#                     finder_Data = finder_Data.append(data, ignore_index=True)
#                     finder_Data = clean(finder_Data, list(finder_Data))

				# if "com.apple.recentitems.plist" in file:
				#     install_Data = ParseInstallHistory(file)
				#     install_Data = clean(install_Data, list(install_Data))
				#     outputToExcel(install_Data, "Installed_software", args.directory, args.project)
# #BROWSER ARTEFACTS
# #-----------------
			if "Chrome" in file:
				if file.endswith("History"):
					owner = re.search(r'([^\\]*)\\[^\\]*$', file).group(1)
					browserData = browser.chrome_history(file)
					browserData.columns = ['URL', 'Title', 'Visit Time']
					browserData['User']=owner
					browserData['Web Browser']="Chrome"
					browserHistory = browserHistory.append(browserData, ignore_index=True)

					browserData = browser.chrome_downloads(file)
					browserData.columns = ['Target Path', 'URL', 'Download Start Time', 'Received Bytes', 'Total Bytes']
					browserData['User']=owner
					browserData['Web Browser']="Chrome"
					browserDownload = browserDownload.append(browserData, ignore_index=True)

			# if "Safari" in file:
			# 	if ("History" in file) and (file.endswith(".webhistory")):
			# 		owner = re.search(r'([^\\]*)\\History', file).group(1)
			# 		browserData = ParseSafariWebHistory(file, owner)
			# 		browserHistory = browserHistory.append(browserData, ignore_index=True)
			# if "Firefox" in file:
			# 	if "places.sqlite" in file:
			# 		owner = re.search(r'Firefox\\([^\\]*)', file).group(1)
			# 		browserData = browser.mozilla_history(file)
			# 		browserData.columns = ['URL', 'Title', 'Visit Time']
			# 		browserData['User Profile']=owner
			# 		browserData['Web Browser']="Mozilla"
			# 		browserHistory = browserHistory.append(browserData, ignore_index=True)


	# outputToExcel(apprecentdocs_Data, "MRU", args.directory, args.project,29,False)
	# outputToExcel(recenthosts_Data, "MRU", args.directory, args.project,23,False)
	# outputToExcel(recentservers_Data, "MRU", args.directory, args.project,17,False)
	# outputToExcel(recentdocs_Data, "MRU", args.directory, args.project,11,False)
	# outputToExcel(recentapp_Data, "MRU", args.directory, args.project,5,False)
	# outputToExcel(finder_Data, "MRU", args.directory, args.project,"",False)
	outputToExcel(browserHistory, "Browser", args.directory, args.project, "", False)
	outputToExcel(browserDownload, "Browser", args.directory, args.project,5, False)