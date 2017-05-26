from config import CONFIG
import os
import argparse
import csv
import logging
import datetime
import psycopg2
import sys
import numpy as num
import pandas as pd
import openpyxl
from sqlalchemy import create_engine
from openpyxl import Workbook, load_workbook
logger = logging.getLogger('root')

def databaseConnect(databaseHost, databaseName, databaseUser, databasePassword):
    databaseConnectionString = "host=" + databaseHost + " dbname=" + databaseName + " user=" + databaseUser + " password=" + databasePassword
    logger.info("databaseConnectionString is " + databaseConnectionString + "\n")
    try:
        databaseConnectionHandle = psycopg2.connect(databaseConnectionString)
    except psycopg2.OperationalError as e:
        logger.error(('Unable to connect!\n{0}').format(e))
        sys.exit(1)
    else:        
        return databaseConnectionHandle

def lsofnetstat(project, directory):
	DATABASE = CONFIG['DATABASE']
	databaseHandle = databaseConnect(DATABASE['HOST'], DATABASE['DATABASENAME'], DATABASE['USER'], DATABASE['PASSWORD'])
	print str(databaseHandle)
	cur = databaseHandle.cursor()

	logger.info("Project is " + project)
	date = str(datetime.datetime.strftime(datetime.datetime.today(),'%Y%m%d%H%M%S'))
		
	#Check if results folder exist, if not, create it.
	dir = os.getcwd()
	resultsDir = dir + "/Results"
	if not os.path.exists(resultsDir):
		try:
			os.makedirs(resultsDir)
		except:
			logging.error("Unable to create results folder")
			sys.exit()

	projResultsDir = dir + "/Results/" + project 
	if not os.path.exists(projResultsDir):
		try:
			os.makedirs(projResultsDir)
		except:
			logging.error("Unable to create Project results folder")
			sys.exit()

	#get hostname
	hostname = os.path.split(directory)[1]
	logger.info("Hostname is " + hostname)

	#Create excel workbook to store results
	workbook = Workbook()
	workbook.create_sheet("Matched", 0)
	workbook.create_sheet("Unmatched", 0)
	destFilename = date + '_' + hostname + '_lsofnetstat-results.xlsx'
	workbook.save(filename='./results/' + project + "/" + destFilename)

	workbook = load_workbook(filename='./results/' + project + "/" + date + '_' + hostname + '_lsofnetstat-results.xlsx')
	matchedrows = workbook["Matched"].max_row
	unmatchedrows = workbook["Unmatched"].max_row
	writer = pd.ExcelWriter('./results/' + project + "/" + date + '_' + hostname + '_lsofnetstat-results.xlsx', engine='openpyxl')
	writer.book = workbook
	writer.sheets = dict((ws.title,ws) for ws in workbook.worksheets)

	for root, dirs, files in os.walk(directory):
		for filename in files:
			#Queueing all triage output files for processing. Once processed, they are removed
			if "lsof-netfiles.txt" in filename:             
				lsofpath = os.path.join(root,filename)
			if "netstat.txt" in filename:             
				netstatpath = os.path.join(root,filename)

	netstat_column_count = 11
	lsof_column_count = 10

	# Next, create the table
	listen_table_name = 'lsoflisten_{0}'.format(date)
	established_table_name = 'lsofestablished_{0}'.format(date)
	netstat_table_name = 'netstat_{0}'.format(date)

	q = 'CREATE TEMPORARY TABLE {0} ('.format(listen_table_name)
	q += ', '.join('col_{0} VARCHAR(255)'.format(i) for i in range(lsof_column_count))
	q += ');'
	cur.execute(q)
	databaseHandle.commit()

	q = 'CREATE TEMPORARY TABLE {0} ('.format(established_table_name)
	q += ', '.join('col_{0} VARCHAR(255)'.format(i) for i in range(lsof_column_count))
	q += ');'
	cur.execute(q)
	databaseHandle.commit()

	q = 'CREATE TEMPORARY TABLE {0} ('.format(netstat_table_name)
	q += ', '.join('col_{0} VARCHAR(255)'.format(i) for i in range(netstat_column_count))
	q += ');'
	cur.execute(q)
	databaseHandle.commit()

	l = "INSERT INTO {0} VALUES (".format(listen_table_name)
	l += ', '.join(('%s ' * lsof_column_count).split())
	l += ');'
	e = "INSERT INTO {0} VALUES (".format(established_table_name)
	e += ', '.join(('%s ' * lsof_column_count).split())
	e += ');'
	n = "INSERT INTO {0} VALUES (".format(netstat_table_name)
	n += ', '.join(('%s ' * netstat_column_count).split())
	n += ');'

	# Populate temp database with lsof entries
	with open(lsofpath) as f:
		reader = csv.reader(f, delimiter=' ')
		next(reader)
		for row in reader:
			row = filter(None, row)
			if "(LISTEN)" in row:
				cur.execute(l, tuple(row))
				databaseHandle.commit()
			if "(ESTABLISHED)" in row:
				cur.execute(e, tuple(row))
				databaseHandle.commit()

	# Populate temp database with netstat entries
	with open(netstatpath) as f:
		reader = csv.reader(f, delimiter=' ')
		next(reader)
		next(reader)
		for row in reader:
			if 'tcp' in row:
				row = filter(None, row)
				cur.execute(n, tuple(row))
				databaseHandle.commit()
			else: 
				if 'udp' in row:
					row = filter(None, row)
					row.insert(5,'')
					cur.execute(n, tuple(row))
					databaseHandle.commit()

	# Select all entries that were populated
	n = "SELECT * FROM {0} ".format(listen_table_name) + ';'
	cur.execute(n)
	lsof_listen = pd.DataFrame(cur.fetchall())

	n = "SELECT * FROM {0} ".format(established_table_name) + ';'
	cur.execute(n)
	lsof_established = pd.DataFrame(cur.fetchall())

	n = "SELECT * FROM {0} ".format(netstat_table_name) + ';'
	cur.execute(n)
	netstat = pd.DataFrame(cur.fetchall())
	netstat['pid'] = netstat[8].str.split("/", expand=True)[0]

	#compare entries
	results = pd.DataFrame()
	#lsof listen match with local address port
	matched = pd.DataFrame()
	unmatched = pd.DataFrame()
	matchednetstat = pd.DataFrame()
	for row in lsof_listen.iterrows():
		pid = row[1][1]
		localport = row[1][8].split(':')[1]
		matched = netstat[netstat['pid'] == pid]
		matched = matched[matched[5] == 'LISTEN']
		matched = matched[matched[3].str.contains(localport, na=False)]
		if len(matched.index) > 0:
			results = results.append(row[1])
			results = results.append(matched)
			matchednetstat = matchednetstat.append(matched)
		else:
			unmatched = unmatched.append(row[1])

	#lsof establish match with local and foreign
	for row in lsof_established.iterrows():
		pid = row[1][1]
		connection = row[1][8].split('->')
		localport = connection[0].split(':')[1]
		foreignport = connection[1].split(':')[1]
		matched = netstat[netstat['pid'] == pid]
		matched = matched[matched[5] == 'ESTABLISHED']
		matched = matched[matched[3].str.contains(localport, na=False) & matched[4].str.contains(foreignport, na=False)]
		if len(matched.index) > 0:
			results = results.append(row[1])
			results = results.append(matched)
			matchednetstat = matchednetstat.append(matched)
		else:
			unmatched = unmatched.append(row[1])

	#output to excel (match tab and unmatch tab) refer to vt output python script
	if not matchedrows == 1:
		results.to_excel(writer, sheet_name="Matched", startrow=matchedrows, header=False, index=False)
	else:
		results.to_excel(writer, sheet_name="Matched", index=False)

	diff = pd.concat([matchednetstat, netstat])
	diff = diff.reset_index(drop=True)
	diff_gpby = diff.groupby(list(diff.columns))
	idx = [x[0] for x in diff_gpby.groups.values() if len(x) == 1]
	diff = diff.reindex(idx)
	unmatched = unmatched.append(diff)

	if not unmatchedrows == 1:
		unmatched.to_excel(writer, sheet_name="Unmatched", startrow=unmatchedrows, header=False, index=False)
	else:
		unmatched.to_excel(writer, sheet_name="Unmatched", index=False)
	writer.save()

def main():
	parser = argparse.ArgumentParser(description="Matches lsof entries with netstat")
	parser.add_argument('-p', dest='projectname', type=str, required=True, help="Codename of the project that the evidence is part of")
	parser.add_argument('-d', dest='directory', type=str, required=True, help="Path to directory containing lsof and netfiles to process")
	
	# parser.add_argument('-lsof', dest='lsof', type=str, required=True, help="Path to lsof file")
	# parser.add_argument('-netstat', dest='netstat', type=str, required=True, help="Path to netstat file")                
	args = parser.parse_args()    
	lsofnetstat(args.projectname, args.directory)
	

if __name__ == '__main__':
	main()
